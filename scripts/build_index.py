"""
Offline index builder for the MedDRA Coding Assistant.

Reads MEDDRA.xlsx and produces two small, deploy-friendly files in ../data/:
  • meddra_terms.jsonl.gz   — one row per LLT: {id, llt, pt, soc}  (~1 MB)
  • meddra_vectors.npz      — int8-quantized semantic embeddings   (~21 MB)

Run this ONCE locally whenever the MedDRA source changes:
    python scripts/build_index.py --xlsx /path/to/MEDDRA.xlsx

The running web service never needs MEDDRA.xlsx — only the two generated files.
"""
import argparse
import gzip
import json
import os
import sys
from collections import Counter

import numpy as np

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, "..", "data")
EMBED_MODEL = "BAAI/bge-small-en-v1.5"


def parse_meddra(xlsx_path):
    import openpyxl

    print(f"Loading workbook (read-only): {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    id_to_name, id_to_type, broader = {}, {}, {}

    print("  reading _ID2NAME ...")
    for r in wb["_ID2NAME"].iter_rows(min_row=2, values_only=True):
        if r[0] is not None:
            id_to_name[int(r[0])] = str(r[2]).strip().strip('"').strip()

    print("  reading _ID2HIERARCHY ...")
    for r in wb["_ID2HIERARCHY"].iter_rows(min_row=2, values_only=True):
        if r[0] is not None:
            id_to_type[int(r[0])] = str(r[2]).strip()

    print("  reading _BROADER ...")
    for r in wb["_BROADER"].iter_rows(min_row=2, values_only=True):
        if r[0] is not None and r[2] is not None:
            broader[int(r[0])] = int(r[2])
    wb.close()

    print("  TYPE COUNTS:", dict(Counter(id_to_type.values())))

    def climb(tid):
        """Walk 'broader' links to find the term's PT and SOC."""
        pt = soc = None
        if id_to_type.get(tid) == "PT":
            pt = id_to_name.get(tid)
        cur, seen, steps = tid, set(), 0
        while cur in broader and steps < 25:
            cur = broader[cur]
            if cur in seen:
                break
            seen.add(cur)
            steps += 1
            t = id_to_type.get(cur)
            if t == "PT" and pt is None:
                pt = id_to_name.get(cur)
            if t == "SOC" and soc is None:
                soc = id_to_name.get(cur)
        return pt, soc

    # LLT rows (the granular synonyms / lay terms) ...
    records = []
    llt_names_lower = set()
    for tid, name in id_to_name.items():
        if id_to_type.get(tid) != "LLT":
            continue
        pt, soc = climb(tid)
        records.append({"id": tid, "llt": name, "pt": pt or name, "soc": soc, "level": "LLT"})
        llt_names_lower.add(name.lower())

    # ... plus PT rows so the clean canonical term (e.g. "Headache", "Drug
    # ineffective") is itself searchable. Skip a PT whose name already exists as
    # an LLT to avoid duplicate vectors.
    pt_added = 0
    for tid, name in id_to_name.items():
        if id_to_type.get(tid) != "PT":
            continue
        if name.lower() in llt_names_lower:
            continue
        _, soc = climb(tid)
        records.append({"id": tid, "llt": name, "pt": name, "soc": soc, "level": "PT"})
        pt_added += 1

    print(f"  Built {len(records)} rows ({len(records)-pt_added} LLT + {pt_added} PT)")
    return records


def embed_terms(texts):
    from fastembed import TextEmbedding

    print(f"Loading embedding model: {EMBED_MODEL}")
    model = TextEmbedding(model_name=EMBED_MODEL)
    print(f"Embedding {len(texts)} terms (this can take ~1-2 min) ...")
    vecs = np.array(list(model.embed(texts, batch_size=256)), dtype=np.float32)
    # L2-normalize so dot product == cosine similarity.
    vecs /= (np.linalg.norm(vecs, axis=1, keepdims=True) + 1e-9)
    return vecs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=os.path.join(HERE, "..", "..", "medguard",
                                                    "medguard_ai-main", "MEDDRA.xlsx"),
                    help="Path to MEDDRA.xlsx")
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: MEDDRA.xlsx not found at {args.xlsx}", file=sys.stderr)
        print("Pass the correct path with --xlsx", file=sys.stderr)
        sys.exit(1)

    os.makedirs(DATA_DIR, exist_ok=True)

    records = parse_meddra(args.xlsx)

    # 1) terms table (compact, gzipped jsonl)
    terms_path = os.path.join(DATA_DIR, "meddra_terms.jsonl.gz")
    with gzip.open(terms_path, "wt", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"Wrote {terms_path}  ({os.path.getsize(terms_path)/1024/1024:.2f} MB)")

    # 2) semantic vectors (int8 quantized). Row order matches the terms file.
    vecs = embed_terms([r["llt"] for r in records])
    q = np.clip(np.round(vecs * 127.0), -127, 127).astype(np.int8)  # normalized -> [-127,127]
    vec_path = os.path.join(DATA_DIR, "meddra_vectors.npz")
    np.savez_compressed(vec_path, vectors=q, model=EMBED_MODEL, dim=vecs.shape[1])
    print(f"Wrote {vec_path}  ({os.path.getsize(vec_path)/1024/1024:.2f} MB)  shape={q.shape}")

    print("\nDone. Copy the two files in data/ with your deploy; MEDDRA.xlsx is NOT needed at runtime.")


if __name__ == "__main__":
    main()
